import openpyxl

book = openpyxl.Workbook()

print(book.sheetnames)

book.create_sheet('meus computadores')

comp_pages = book['meus computadores']
comp_pages.append(['Computador 1','8 GB Ram','R$ 2,500'])
comp_pages.append(['Computador 2','16 GB Ram','R$ 5,500'])
comp_pages.append(['Computador 3','32 GB Ram','R$ 8,500'])

book.save('meus computadores.xlsx')