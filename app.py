import openpyxl

# Criar uma planilha(book).
book = openpyxl.Workbook()
# Como visualizar páginas existentes.
print(book.sheetnames)
# Como criar uma página.
book.create_sheet('Frutas')
# Como selecionar uma página.
frutas_page = book['Frutas']
frutas_page.append(['Frutas', 'Quantidade', 'Preço'])
frutas_page.append(['Banana', '5', 'R$3,90'])
frutas_page.append(['Ameixa', '8', 'R$2,00'])
frutas_page.append(['Melancia', '2', 'R$3,90'])
frutas_page.append(['Morango', '1', 'R$4,30'])
# Salvar a planilha.
book.save('Planilha de Compras.xlsx')
